"""
Reads Bloomberg Excel template, exports to CSV, and runs the dashboard.

Usage:
    1. Open bloomberg_template.xlsx in Excel with Bloomberg Add-in
    2. Wait for data to populate
    3. Save the file (Ctrl+S) to cache the values
    4. Run this script: python scripts/run_dashboard.py
"""

import sys
from pathlib import Path

# Add src to path
sys.path.insert(0, str(Path(__file__).parent.parent / "src"))

import openpyxl
import csv
from datetime import datetime


def export_bloomberg_to_csv():
    """Read Bloomberg Excel and export to CSV."""

    template_path = Path(__file__).parent.parent / "data" / "exports" / "bloomberg_template.xlsx"
    csv_path = Path(__file__).parent.parent / "data" / "exports" / "bloomberg_data.csv"

    if not template_path.exists():
        print(f"ERROR: Bloomberg template not found: {template_path}")
        return False

    print(f"Reading: {template_path}")

    # Load workbook with data_only=True to get cached values instead of formulas
    wb = openpyxl.load_workbook(template_path, data_only=True)

    # Use the Export sheet (has clean tickers)
    if "Export" in wb.sheetnames:
        ws = wb["Export"]
        print("Using 'Export' sheet")
    else:
        ws = wb.active
        print(f"Using active sheet: {ws.title}")

    # Get headers from row 1
    headers = []
    for cell in ws[1]:
        if cell.value:
            headers.append(cell.value)
        else:
            break

    print(f"Columns: {headers}")

    # Export to CSV
    rows_exported = 0
    with open(csv_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(headers)

        for row in ws.iter_rows(min_row=2, max_col=len(headers)):
            # Skip empty rows
            if not row[0].value:
                continue

            values = []
            for cell in row:
                val = cell.value
                # Handle None and format numbers
                if val is None:
                    values.append("")
                elif isinstance(val, float):
                    values.append(f"{val}")
                else:
                    values.append(str(val))

            writer.writerow(values)
            rows_exported += 1

    print(f"Exported {rows_exported} rows to: {csv_path}")
    wb.close()
    return True


def run_dashboard():
    """Run the main dashboard report generator."""

    print("\n" + "="*60)
    print("Running DAT Dashboard...")
    print("="*60 + "\n")

    # Import and run main
    try:
        from main import main
        main()
    except ImportError as e:
        print(f"Error importing main: {e}")
        print("Trying alternative import...")

        import subprocess
        main_path = Path(__file__).parent.parent / "src" / "main.py"
        result = subprocess.run([sys.executable, str(main_path)], cwd=str(main_path.parent.parent))
        return result.returncode == 0

    return True


def main():
    print("="*60)
    print("DAT Dashboard - Bloomberg Export & Run")
    print(f"Time: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("="*60)
    print()

    # Step 1: Export Bloomberg data to CSV
    print("Step 1: Exporting Bloomberg data to CSV...")
    if not export_bloomberg_to_csv():
        print("Failed to export Bloomberg data.")
        return 1

    # Step 2: Run dashboard
    print("\nStep 2: Running dashboard report...")
    if not run_dashboard():
        print("Failed to run dashboard.")
        return 1

    print("\n" + "="*60)
    print("Done!")
    print("="*60)
    return 0


if __name__ == "__main__":
    sys.exit(main())
