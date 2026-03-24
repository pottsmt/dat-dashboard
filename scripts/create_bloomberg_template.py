"""Creates the Bloomberg Excel template with per-ticker VWAP dates."""

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def create_template():
    wb = Workbook()

    # Data sheet
    data_sheet = wb.active
    data_sheet.title = 'Data'

    # Headers - added vwap_start_date and vwap columns
    headers = [
        'ticker', 'price', 'volume', 'avg_vol_5d', 'shares_out', 'market_cap',
        'total_debt', 'cash', 'enterprise_value', 'vwap_start_date', 'vwap',
        'price_3d_ago', 'price_7d_ago', 'price_30d_ago'
    ]

    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='2D3748', end_color='2D3748', fill_type='solid')
    date_fill = PatternFill(start_color='E6F3FF', end_color='E6F3FF', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    for col, header in enumerate(headers, 1):
        cell = data_sheet.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    # Tickers with default VWAP start dates (update these to actual last filing dates)
    tickers = [
        ('ALTS US Equity', '2025-12-01'),
        ('ASST US Equity', '2025-12-15'),
        ('AVX US Equity', '2025-12-01'),
        ('BMNR US Equity', '2025-12-15'),
        ('BNC US Equity', '2025-12-01'),
        ('BRR US Equity', '2025-12-01'),
        ('BTBT US Equity', '2025-12-15'),
        ('BTOG US Equity', '2025-09-01'),
        ('CEPO US Equity', '2025-12-01'),
        ('CYPH US Equity', '2025-12-01'),
        ('DFDV US Equity', '2025-12-15'),
        ('EMPD US Equity', '2025-12-01'),
        ('ETHM US Equity', '2025-12-01'),
        ('ETHZ US Equity', '2025-12-01'),
        ('FGNX US Equity', '2025-12-01'),
        ('FWDI US Equity', '2025-12-01'),
        ('GNLN US Equity', '2025-12-01'),
        ('HSDT US Equity', '2025-12-01'),
        ('HYPD US Equity', '2025-12-01'),
        ('IPST US Equity', '2025-09-01'),
        ('LITS US Equity', '2025-09-01'),
        ('MLAC US Equity', '2025-12-01'),
        ('MSTR US Equity', '2026-01-15'),
        ('NA US Equity', '2025-12-01'),
        ('NAKA US Equity', '2025-09-01'),
        ('ORBS US Equity', '2025-12-01'),
        ('PAPL US Equity', '2025-12-01'),
        ('PURR US Equity', '2025-12-01'),
        ('SBET US Equity', '2025-12-01'),
        ('SLMT US Equity', '2025-09-01'),
        ('STSS US Equity', '2025-12-01'),
        ('SUIG US Equity', '2025-12-01'),
        ('SVRN US Equity', '2025-09-01'),
        ('THAR US Equity', '2025-09-01'),
        ('TLGYF US Equity', '2025-09-01'),
        ('TONX US Equity', '2025-12-01'),
        ('TRON US Equity', '2025-12-01'),
        ('UPXI US Equity', '2025-12-01'),
        ('ZONE US Equity', '2025-09-01'),
    ]

    for row, (ticker, vwap_date) in enumerate(tickers, 2):
        # Ticker
        data_sheet.cell(row=row, column=1, value=ticker).border = thin_border
        # Price
        data_sheet.cell(row=row, column=2, value=f'=BDP($A{row},"PX_LAST")').border = thin_border
        # Volume
        data_sheet.cell(row=row, column=3, value=f'=BDP($A{row},"VOLUME")').border = thin_border
        # 5D Avg Volume
        data_sheet.cell(row=row, column=4, value=f'=BDP($A{row},"VOLUME_AVG_5D")').border = thin_border
        # Shares Outstanding
        data_sheet.cell(row=row, column=5, value=f'=BDP($A{row},"EQY_SH_OUT")').border = thin_border
        # Market Cap
        data_sheet.cell(row=row, column=6, value=f'=BDP($A{row},"CUR_MKT_CAP")').border = thin_border
        # Total Debt
        data_sheet.cell(row=row, column=7, value=f'=BDP($A{row},"BS_TOT_LIAB2")').border = thin_border
        # Cash
        data_sheet.cell(row=row, column=8, value=f'=BDP($A{row},"BS_CASH_NEAR_CASH_ITEM")').border = thin_border
        # Enterprise Value
        data_sheet.cell(row=row, column=9, value=f'=BDP($A{row},"CURR_ENTP_VAL")').border = thin_border

        # VWAP start date - editable per ticker (highlighted blue)
        vwap_cell = data_sheet.cell(row=row, column=10, value=vwap_date)
        vwap_cell.border = thin_border
        vwap_cell.fill = date_fill

        # INTERVAL_VWAP - multi-day VWAP using date overrides
        vwap_formula = f'=BDP($A{row},"INTERVAL_VWAP","START_DATE_OVERRIDE",$J{row},"END_DATE_OVERRIDE",TODAY()-1)'
        data_sheet.cell(row=row, column=11, value=vwap_formula).border = thin_border

        # Historical prices - use INDEX to extract single value from BDH result
        # BDH returns [date, price] so INDEX(...,1,2) gets the price
        # Use calendar days that roughly equal trading days: 5->3td, 10->7td, 45->30td
        data_sheet.cell(row=row, column=12, value=f'=INDEX(BDH($A{row},"PX_LAST",TODAY()-5,TODAY()-5,"Fill","P"),1,2)').border = thin_border
        data_sheet.cell(row=row, column=13, value=f'=INDEX(BDH($A{row},"PX_LAST",TODAY()-10,TODAY()-10,"Fill","P"),1,2)').border = thin_border
        data_sheet.cell(row=row, column=14, value=f'=INDEX(BDH($A{row},"PX_LAST",TODAY()-45,TODAY()-45,"Fill","P"),1,2)').border = thin_border

    # Set column widths
    col_widths = [18, 12, 12, 12, 14, 14, 14, 14, 16, 14, 12, 12, 12, 12]
    for i, width in enumerate(col_widths, 1):
        data_sheet.column_dimensions[get_column_letter(i)].width = width

    # Export sheet - clean data for CSV
    export_sheet = wb.create_sheet('Export')

    for col, header in enumerate(headers, 1):
        cell = export_sheet.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    for row in range(2, len(tickers) + 2):
        # Clean ticker (extract just symbol from "MSTR US Equity")
        export_sheet.cell(row=row, column=1, value=f'=LEFT(Data!A{row},FIND(" ",Data!A{row})-1)').border = thin_border
        # Reference other columns from Data sheet
        for col in range(2, 15):
            export_sheet.cell(row=row, column=col, value=f'=Data!{get_column_letter(col)}{row}').border = thin_border

    for i, width in enumerate(col_widths, 1):
        export_sheet.column_dimensions[get_column_letter(i)].width = width

    # Instructions sheet
    instructions = wb.create_sheet('Instructions')
    instructions['A1'] = 'DAT Dashboard - Bloomberg Template'
    instructions['A1'].font = Font(bold=True, size=16)

    instructions['A3'] = 'How to use:'
    instructions['A3'].font = Font(bold=True)
    instructions['A4'] = '1. Enable Bloomberg Add-in (Bloomberg menu > Start Bloomberg)'
    instructions['A5'] = '2. Data sheet auto-populates with live data'
    instructions['A6'] = '3. Column J (vwap_start_date) - SET THIS per ticker to last shares filing date'
    instructions['A6'].font = Font(bold=True, color='0000FF')
    instructions['A7'] = '4. Column K (vwap) calculates VWAP from that date to yesterday'
    instructions['A8'] = '5. Export sheet has clean tickers for CSV export'

    instructions['A10'] = 'VWAP Date Column (J) - IMPORTANT:'
    instructions['A10'].font = Font(bold=True)
    instructions['A11'] = 'Set this to the date of the last 10-Q, 10-K, or 8-K that reported shares outstanding.'
    instructions['A12'] = 'The VWAP from this date is used to estimate ATM dilution since that filing.'
    instructions['A13'] = 'Example: If MSTR filed 10-K on Jan 15 with shares count, set J row to 2026-01-15'

    instructions['A15'] = 'VBA Export Macro (Alt+F11 > Insert Module):'
    instructions['A15'].font = Font(bold=True)
    instructions['A17'] = 'Sub ExportToCSV()'
    instructions['A18'] = '    Sheets("Export").Copy'
    instructions['A19'] = '    ActiveWorkbook.SaveAs "C:/Users/Matthew/claude-projects/dat-dashboard/data/exports/bloomberg_data.csv", xlCSV'
    instructions['A20'] = '    ActiveWorkbook.Close False'
    instructions['A21'] = 'End Sub'

    instructions.column_dimensions['A'].width = 100

    # Save
    output_path = 'data/exports/bloomberg_template.xlsx'
    wb.save(output_path)
    print(f'Created: {output_path}')
    print('')
    print('Key columns:')
    print('  J: vwap_start_date - EDIT per ticker (blue cells) to last shares filing date')
    print('  K: vwap - Auto-calculated from start date to yesterday')
    print('')
    print('The VWAP is used to estimate shares sold via ATM between filings.')


if __name__ == '__main__':
    create_template()
